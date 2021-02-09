# import pyexcel as p
# FULL_PATH1 = r'C:\Users\Rustem\PycharmProjects\python3\example.XLS'
# FULL_OUT_PATH = r'C:\Users\Rustem\PycharmProjects\python3\example.xlsx'
# p.save_book_as(file_name=FULL_PATH1, dest_file_name=FULL_OUT_PATH, font=)

import openpyxl as op
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
import os
import re


import logging
import sys

# LOGGER_FORMAT = u'%(filename)s[LINE:%(lineno)d]# %(levelname)-8s [%(asctime)s]  %(message)s'
# LOGGER_HANDLERS = [logging.FileHandler("debug.log"), logging.StreamHandler(sys.stdout)]
LOGGER_FORMAT = u'%(message)s'
LOGGER_HANDLERS = [logging.StreamHandler(sys.stdout)]
logging.basicConfig(
    format=LOGGER_FORMAT,
    level=logging.INFO,
    handlers=LOGGER_HANDLERS)

LOGGER = logging.getLogger("LOGGER")


SOURCE_PATH = r'/python3/Budget_script/source/'


class RowFormatStyle:

    @staticmethod
    def new(name, regex_pattern, fill, font_size):
        result = RowFormatStyle()
        result.name = name
        result.regex_pattern = regex_pattern
        result.fill = PatternFill(start_color=fill, fill_type='solid')
        result.font = Font(bold=True, size=font_size)
        return result

    name: str
    # Регулярное выражение для парсинга кода дохода. Определяет уровень.
    regex_pattern: str
    fill: PatternFill
    font: Font


LEVEL1_FORMAT = RowFormatStyle.new("Level1", r"^[1-2]\|00", "FFFF00", 11)
LEVEL2_FORMAT = RowFormatStyle.new("Level2", r"^[1-2]\|[0-9][1-9]\|00", "FFCC00", 10)
LEVEL3_FORMAT = RowFormatStyle.new("Level3", r"^[1-2]\|[0-9][1-9]\|\d{2}\|0{3}\|0[0-2]", "E6BBC1", 9)
LEVEL4_FORMAT = RowFormatStyle.new("Level4", r"^[1-2]\|[0-9][1-9]\|\d{2}\|[0-9]{2}0\|", "DCDCDC", 8)

PATTERNS = [LEVEL1_FORMAT, LEVEL2_FORMAT, LEVEL3_FORMAT, LEVEL4_FORMAT]


def split_tax_code(code):
    """
    Метод разбивает код дохода на массив отдельных кодов классификации
    :param code: код дохода
    :return: list
    """
    return tuple((code[0], code[1:3], code[3:5], code[5:8], code[8:10], code[10:14], code[14:]))


def refactor_tax_code(code):
    """
    Метод форматирует код дохода по заданному шаблону
    :param code: код дохода
    :return: str
    """
    return "%s|%s|%s|%s|%s|%s|%s" % split_tax_code(code)


def find_last_not_null_row(sheet, start_row, column):
    """
    Для указанного стролбца метод находит номер последней строки, содержащей данные
    т.е. последней НЕпустой строки
    :param sheet:
    :param start_row: строка, с которой начнётся поиск
    :param column: столбец, в котором будет проводиться поиск
    :return: int
    """
    cell_value = sheet.cell(row=start_row, column=column).value
    i = 0
    while cell_value is not None:
        i = i + 1
        cell_value = sheet.cell(row=start_row + i, column=column).value

    return start_row + i - 1


def format_cell_list(cell_list, font=None, fill=None, pattern: RowFormatStyle = None):
    # Форматирование списка ячеек.
    # Применение шрифта, заливки, либо параметров RowFormatStyle
    if pattern is not None:
        font = pattern.font
        fill = pattern.fill

    if font is not None:
        for _cell in cell_list:
            _cell.font = font

    if fill is not None:
        for _cell in cell_list:
            _cell.fill = fill


def replace_substr_in_all_cells(sheet, last_row, replace_substr, to_replace_substr):
    # Проход по всем ячейкам и замена исходной строки на заданную
    for row in range(1, last_row + 1):
        for col in range(1, 35):
            cell_val = sheet.cell(row=row, column=col).value
            if cell_val is None:
                continue
            if not isinstance(cell_val, str):
                continue
            if replace_substr in cell_val:
                sheet.cell(row=row, column=col).value = \
                    cell_val.replace(replace_substr, to_replace_substr)


def format_tax_codes_in_all_rows(sheet, column, start_row, last_row):
    # Форматирование кодоы дохода и отдельные кодов классификации во всех строках
    # Форматирование ширины колонок.
    # Работает для одной из двух страниц на листе
    def set_text_with_alignment(col, text):
        cell = sheet.cell(row=14, column=col)
        cell.value = text
        cell.font = Font(name="Arial", bold=True, size=9)
        cell.alignment = Alignment(wrapText=True, textRotation=90)

    for row in range(start_row, last_row + 1):
        tax_full_code = sheet.cell(row=row, column=4).value.replace("'", "")
        tax_full_code_array = split_tax_code(tax_full_code)
        tax_full_code_refactored = refactor_tax_code(tax_full_code)
        # в отдельные колонки вставляются коды классификации
        for i in range(0, 5):
            sheet.cell(row=row, column=column + i).value = tax_full_code_array[i]
            sheet.column_dimensions[get_column_letter(column + i)].width = 4

        # исходный код дохода заменяется на отформатированный
        sheet.cell(row=row, column=column-1).value = tax_full_code_refactored

    set_text_with_alignment(column, "Группа доходов")
    set_text_with_alignment(column + 1, "Подгруппа доходов")
    set_text_with_alignment(column + 2, "Статья доходов")
    set_text_with_alignment(column + 3, "Подстатья доходов")
    set_text_with_alignment(column + 4, "Элемент доходов")


def format_all_rows(sheet, start_row, last_row):
    # Проход по строкам, анализ кода дохода по регулярным выражениям,
    # определение уровня и форматирование строки в соответствии с этим уровнем
    sheet_rows_list = list(sheet.rows)
    for row_index in range(start_row, last_row + 1):
        tax_full_code = sheet.cell(row=row_index, column=4).value
        for pattern in PATTERNS:
            if re.match(pattern.regex_pattern, tax_full_code):
                format_cell_list(sheet_rows_list[row_index - 1], pattern=pattern)
                break


def format_all_cells_with_numbers(sheet, last_row, start_col, end_col):
    # Форматирование числового формат в заданном диапазоне ячеек
    for row in range(16, last_row + 1):
        for col in range(start_col, end_col + 1):
            sheet.cell(row=row, column=col).number_format = "### ### ### ### ##0"


def hide_some_columns(sheet, last_row):
    # Сокрытие колонок. Работает для всех (двух) страниц на листе.
    for col in range(2, 44):
        if col == 2 or col in range(10, 32):
            sheet.column_dimensions[get_column_letter(col)].hidden = True
            continue
        need_to_hide = True
        for row in range(16, last_row + 1):
            if sheet.cell(row=row, column=col).value is not None:
                need_to_hide = False
                break
        if need_to_hide:
            sheet.column_dimensions[get_column_letter(col)].hidden = True


def format_columns_width(sheet, column):
    # Форматирование ширины колонок.
    # Работает для одной из двух страниц на листе.
    sheet.column_dimensions[get_column_letter(column)].width = 50
    sheet.column_dimensions[get_column_letter(column + 2)].width = 4
    sheet.column_dimensions[get_column_letter(column + 3)].width = 22
    for i in range(9, 22):
        sheet.column_dimensions[get_column_letter(column + i)].width = 17


def refactor_sheet1(sheet):
    # Вставка по 5 колонок в каждую из двух страниц на листе
    # сначала на 2-ой странице - колонки 22-26
    # потом на 1-ой странице - колонки 5-9
    for i in range(0, 2):
        for j in range(0, 5):
            sheet.insert_cols(22 - 17*i)

    sheet_rows_list = list(sheet.rows)
    sheet_cols_list = list(sheet.columns)

    # Включение переноса текста во всех ячейках в 1-ой колонке
    for cell in sheet_cols_list[0]:
        cell.alignment = Alignment(wrapText=True)

    sheet.row_dimensions[14].height = 100

    # Проставление в 10 строке номеров колонок
    sheet.row_dimensions[10].height = 30
    for cell in sheet_rows_list[9]:
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(name="Arial", size=16)
        cell.value = cell.column

    last_row = find_last_not_null_row(sheet, 16, 1)

    replace_substr_in_all_cells(sheet, last_row, "Российской Федерации", "РФ")

    format_cell_list(sheet_rows_list[15], font=Font(bold=True, size=10),
                     fill=PatternFill(start_color='92d050', fill_type='solid'))

    # Вызов для 1-ой страницы
    format_tax_codes_in_all_rows(sheet, 5, 18, last_row)
    # Вызов для 2-ой страницы
    format_tax_codes_in_all_rows(sheet, 27, 18, last_row)

    # Вызов для всего листа
    format_all_rows(sheet, 18, last_row)

    # Вызов для 1-ой страницы
    format_all_cells_with_numbers(sheet, last_row, 10, 22)
    # Вызов для 2-ой страницы
    format_all_cells_with_numbers(sheet, last_row, 32, 44)

    # Вызов для 1-ой страницы
    format_columns_width(sheet, 1)
    # Вызов для 2-ой страницы
    format_columns_width(sheet, 23)

    # вызов для всего листа
    hide_some_columns(sheet, last_row)


def main():
    source_files_list = list()
    if os.path.isfile(SOURCE_PATH):
        source_files_list.append(SOURCE_PATH)
    else:
        for file in os.listdir(SOURCE_PATH):
            source_files_list.append(SOURCE_PATH + file)

    for file_path in source_files_list:
        dest_full_path = file_path.replace(".xlsx", "") + "_ref.xlsx"
        wb = op.load_workbook(filename=file_path)
        sheet1 = wb['стр.1_2']
        refactor_sheet1(sheet1)
        wb.save(dest_full_path)
        logging.info("Handled file %s " % dest_full_path)


main()
